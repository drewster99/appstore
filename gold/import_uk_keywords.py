#!/usr/bin/env python3
"""Import UK keywords from existing Excel file into report 1."""

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed")
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).parent))
from db.database import transaction

def score_rank_in_genre(rank: int) -> int:
    if 1 <= rank <= 10: return 3
    elif 11 <= rank <= 25: return 2
    elif 26 <= rank <= 50: return 1
    return 0

def score_popularity_in_genre(popularity: int) -> int:
    if 76 <= popularity <= 100: return 3
    elif 61 <= popularity <= 75: return 2
    elif 50 <= popularity <= 60: return 1
    return 0

def score_overall_popularity(popularity: int) -> int:
    if 86 <= popularity <= 100: return 5
    elif 71 <= popularity <= 85: return 4
    elif 61 <= popularity <= 70: return 3
    elif 51 <= popularity <= 60: return 2
    elif 41 <= popularity <= 50: return 1
    return 0

excel_path = Path("../gold/Month,_Country_or_Region,_Genre,_Search_Term,_Rank_in_Genre,_Search_Popularity_in_Genre_(1-100),_Sea-2.xlsx")

print(f"Loading {excel_path.name}...")
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb.active

# Find header row
headers = list(sheet.iter_rows(min_row=7, max_row=7, values_only=True))[0]
month_idx = headers.index("Month")
country_idx = headers.index("Country or Region")
genre_idx = headers.index("Genre")
term_idx = headers.index("Search Term")
rank_idx = headers.index("Rank in Genre")
pop_genre_idx = headers.index("Search Popularity in Genre (1-100)")
pop_overall_idx = headers.index("Search Popularity (1-100)")
pop_scale_idx = headers.index("Search Popularity (1-5)")

# Collect UK keywords
keywords = []
for i, row in enumerate(sheet.iter_rows(min_row=8, values_only=True), start=8):
    if not row[term_idx]:
        continue

    country = row[country_idx]
    if country != "United Kingdom":
        continue

    rank = int(row[rank_idx]) if row[rank_idx] else 0
    pop_genre = int(row[pop_genre_idx]) if row[pop_genre_idx] else 0
    pop_overall = int(row[pop_overall_idx]) if row[pop_overall_idx] else 0
    pop_scale = int(row[pop_scale_idx]) if row[pop_scale_idx] else 0

    score_rank = score_rank_in_genre(rank)
    score_genre = score_popularity_in_genre(pop_genre)
    score_overall = score_overall_popularity(pop_overall)
    total_score = score_rank + score_genre + score_overall

    keywords.append((
        1,  # report_id
        country,
        row[genre_idx],
        row[term_idx],
        rank,
        pop_genre,
        pop_overall,
        pop_scale,
        score_rank,
        score_genre,
        score_overall,
        total_score
    ))

wb.close()

print(f"Found {len(keywords)} UK keywords")
print("Inserting into database...")

with transaction() as conn:
    conn.executemany("""
        INSERT INTO apple_keywords (
            report_id, country, genre, search_term, rank_in_genre,
            popularity_genre, popularity_overall, popularity_scale,
            score_rank, score_genre, score_overall, total_score
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, keywords)

# Update report total
with transaction() as conn:
    conn.execute("""
        UPDATE apple_reports
        SET total_keywords = (SELECT COUNT(*) FROM apple_keywords WHERE report_id = 1)
        WHERE id = 1
    """)

print(f"✓ Successfully imported {len(keywords)} UK keywords to report 1")
