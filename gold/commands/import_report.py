#!/usr/bin/env python3
"""Import Apple Search Ads Monthly Keyword Rankings report into database."""

import sys
from pathlib import Path
from datetime import datetime
from typing import Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from db.database import transaction, execute_one, execute_many


def score_rank_in_genre(rank: int) -> int:
    """Score based on Rank in Genre position."""
    if 1 <= rank <= 10:
        return 3
    elif 11 <= rank <= 25:
        return 2
    elif 26 <= rank <= 50:
        return 1
    return 0


def score_popularity_in_genre(popularity: int) -> int:
    """Score based on Search Popularity in Genre (1-100)."""
    if 76 <= popularity <= 100:
        return 3
    elif 61 <= popularity <= 75:
        return 2
    elif 50 <= popularity <= 60:
        return 1
    return 0


def score_overall_popularity(popularity: int) -> int:
    """Score based on overall Search Popularity (1-100)."""
    if 86 <= popularity <= 100:
        return 5
    elif 71 <= popularity <= 85:
        return 4
    elif 61 <= popularity <= 70:
        return 3
    elif 50 <= popularity <= 60:
        return 2
    return 0


def parse_excel_metadata(sheet) -> Tuple[str, str, str, str]:
    """
    Parse metadata from Excel file header.

    Returns:
        (report_id, generated_at, data_month, user_locale)
    """
    # Row 1: "Data extract produced by 93070_144880 on 10/13/2025 17:50"
    row1 = str(sheet.cell(1, 1).value or "")

    # Extract report ID and timestamp
    report_id = None
    generated_at = None

    if "produced by" in row1 and "on" in row1:
        parts = row1.split("produced by")[1].split("on")
        if len(parts) == 2:
            report_id = parts[0].strip()
            timestamp_str = parts[1].strip()
            # Parse timestamp: "10/13/2025 17:50"
            try:
                generated_at = datetime.strptime(timestamp_str, "%m/%d/%Y %H:%M")
            except ValueError:
                pass

    # Row 3: "Month = 2025-09"
    row3 = str(sheet.cell(3, 1).value or "")
    data_month = None
    if "Month =" in row3:
        data_month = row3.split("=")[1].strip()

    # Row 5: "userLocale = en_US"
    row5 = str(sheet.cell(5, 1).value or "")
    user_locale = None
    if "userLocale =" in row5:
        user_locale = row5.split("=")[1].strip()

    # Validate we got all required fields
    if not all([report_id, generated_at, data_month, user_locale]):
        raise ValueError(
            f"Could not parse all metadata from Excel file. "
            f"Got: report_id={report_id}, generated_at={generated_at}, "
            f"data_month={data_month}, user_locale={user_locale}"
        )

    return report_id, generated_at.isoformat(), data_month, user_locale


def check_existing_report(report_id: str, generated_at: str) -> Optional[int]:
    """
    Check if this exact report already exists.

    Returns:
        Report ID if exists, None otherwise
    """
    row = execute_one(
        "SELECT id FROM apple_reports WHERE report_id = ? AND generated_at = ?",
        (report_id, generated_at)
    )
    return row['id'] if row else None


def deactivate_old_reports(month_locale_key: str, except_id: Optional[int] = None):
    """
    Deactivate all reports for this month+locale except the specified one.

    Args:
        month_locale_key: The month_locale_key to deactivate
        except_id: Report ID to keep active (if any)
    """
    with transaction() as conn:
        if except_id:
            conn.execute(
                """UPDATE apple_reports
                   SET is_active = 0
                   WHERE month_locale_key = ? AND id != ?""",
                (month_locale_key, except_id)
            )
        else:
            conn.execute(
                "UPDATE apple_reports SET is_active = 0 WHERE month_locale_key = ?",
                (month_locale_key,)
            )


def import_report(
    excel_path: Path,
    country_filter: str = "United States",
    verbose: bool = True
) -> int:
    """
    Import Apple Search Ads report from Excel file.

    Args:
        excel_path: Path to the .xlsx file
        country_filter: Only import keywords for this country
        verbose: Print progress messages

    Returns:
        Report ID of the imported report
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"File not found: {excel_path}")

    if verbose:
        print(f"Parsing Excel file: {excel_path.name}", file=sys.stderr)

    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active

    # Parse metadata
    report_id, generated_at, data_month, user_locale = parse_excel_metadata(sheet)
    month_locale_key = f"{data_month}_{user_locale}"

    if verbose:
        print(f"  Report ID: {report_id}", file=sys.stderr)
        print(f"  Generated: {generated_at}", file=sys.stderr)
        print(f"  Data Month: {data_month}", file=sys.stderr)
        print(f"  Locale: {user_locale}", file=sys.stderr)

    # Check if already imported
    existing_id = check_existing_report(report_id, generated_at)
    if existing_id:
        if verbose:
            print(f"  ℹ Report already imported (ID: {existing_id})", file=sys.stderr)
        wb.close()
        return existing_id

    # Find header row (row 7)
    headers = None
    for i, row in enumerate(sheet.iter_rows(min_row=7, max_row=7, values_only=True), start=7):
        if row and row[0] == "Month":
            headers = list(row)
            break

    if not headers:
        raise ValueError("Could not find valid header row at row 7")

    # Find column indices
    try:
        month_idx = headers.index("Month")
        country_idx = headers.index("Country or Region")
        genre_idx = headers.index("Genre")
        term_idx = headers.index("Search Term")
        rank_idx = headers.index("Rank in Genre")
        pop_genre_idx = headers.index("Search Popularity in Genre (1-100)")
        pop_overall_idx = headers.index("Search Popularity (1-100)")
        pop_scale_idx = headers.index("Search Popularity (1-5)")
    except ValueError as e:
        raise ValueError(f"Could not find required column. Headers: {headers}") from e

    # Collect keywords
    if verbose:
        print(f"Importing keywords for country: {country_filter}...", file=sys.stderr)

    keywords = []
    processed_count = 0
    skipped_count = 0

    for row in sheet.iter_rows(min_row=8, values_only=True):
        processed_count += 1

        # Progress indicator
        if verbose and processed_count % 10000 == 0:
            print(f"  Processed {processed_count:,} rows...", file=sys.stderr)

        if not row[term_idx]:  # Skip empty search terms
            skipped_count += 1
            continue

        # Filter by country
        country = row[country_idx]
        if country_filter and country != country_filter:
            skipped_count += 1
            continue

        try:
            rank = int(row[rank_idx]) if row[rank_idx] else 999
            pop_genre = int(row[pop_genre_idx]) if row[pop_genre_idx] else 0
            pop_overall = int(row[pop_overall_idx]) if row[pop_overall_idx] else 0
            pop_scale = int(row[pop_scale_idx]) if row[pop_scale_idx] else 0
        except (ValueError, TypeError):
            skipped_count += 1
            continue

        # Calculate scores
        rank_score = score_rank_in_genre(rank)
        genre_score = score_popularity_in_genre(pop_genre)
        overall_score = score_overall_popularity(pop_overall)
        total_score = rank_score + genre_score + overall_score

        keywords.append({
            'country': country,
            'genre': str(row[genre_idx]) if row[genre_idx] else "",
            'search_term': str(row[term_idx]),
            'rank_in_genre': rank,
            'popularity_genre': pop_genre,
            'popularity_overall': pop_overall,
            'popularity_scale': pop_scale,
            'score_rank': rank_score,
            'score_genre': genre_score,
            'score_overall': overall_score,
            'total_score': total_score
        })

    wb.close()

    total_keywords = len(keywords)

    if verbose:
        print(f"  Processed {processed_count:,} rows, skipped {skipped_count:,} rows", file=sys.stderr)
        print(f"  Importing {total_keywords:,} keywords into database...", file=sys.stderr)

    # Import into database in a transaction
    with transaction() as conn:
        # Deactivate old reports for this month+locale
        deactivate_old_reports(month_locale_key)

        # Insert report
        cursor = conn.execute(
            """INSERT INTO apple_reports
               (report_id, generated_at, data_month, user_locale, month_locale_key,
                source_filename, total_keywords, is_active)
               VALUES (?, ?, ?, ?, ?, ?, ?, 1)""",
            (report_id, generated_at, data_month, user_locale, month_locale_key,
             excel_path.name, total_keywords)
        )
        new_report_id = cursor.lastrowid

        # Bulk insert keywords
        keyword_rows = [
            (new_report_id, kw['country'], kw['genre'], kw['search_term'],
             kw['rank_in_genre'], kw['popularity_genre'], kw['popularity_overall'],
             kw['popularity_scale'], kw['score_rank'], kw['score_genre'],
             kw['score_overall'], kw['total_score'])
            for kw in keywords
        ]

        conn.executemany(
            """INSERT INTO apple_keywords
               (report_id, country, genre, search_term, rank_in_genre,
                popularity_genre, popularity_overall, popularity_scale,
                score_rank, score_genre, score_overall, total_score)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            keyword_rows
        )

    if verbose:
        print(f"✓ Imported {total_keywords:,} keywords", file=sys.stderr)
        print(f"✓ Set as active report for {month_locale_key}", file=sys.stderr)

    return new_report_id


def main():
    """Command-line interface for importing reports."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Import Apple Search Ads Monthly Keyword Rankings report"
    )
    parser.add_argument(
        "excel_file",
        help="Path to the .xlsx file"
    )
    parser.add_argument(
        "--country",
        default="United States",
        help="Country to filter for (default: United States)"
    )
    parser.add_argument(
        "-q", "--quiet",
        action="store_true",
        help="Suppress progress output"
    )

    args = parser.parse_args()

    try:
        excel_path = Path(args.excel_file)
        report_id = import_report(
            excel_path,
            country_filter=args.country,
            verbose=not args.quiet
        )

        if not args.quiet:
            print(f"\n✓ Import complete! Report ID: {report_id}", file=sys.stderr)

        sys.exit(0)

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
