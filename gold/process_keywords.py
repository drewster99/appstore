#!/usr/bin/env python3
"""
Process Apple Search Ads Monthly Keyword Rankings and score them.
Outputs JSON with scored keywords ready for HTML generation.

Now reads from database by default. Use --from-excel to read from Excel file.
"""

import sys
import json
from pathlib import Path

# Add current directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))

from db.database import execute_one, execute_query

try:
    import openpyxl
except ImportError:
    openpyxl = None


def score_rank_in_genre(rank):
    """Score based on Rank in Genre position."""
    if 1 <= rank <= 10:
        return 3
    elif 11 <= rank <= 25:
        return 2
    elif 26 <= rank <= 50:
        return 1
    return 0


def score_popularity_in_genre(popularity):
    """Score based on Search Popularity in Genre (1-100)."""
    if 76 <= popularity <= 100:
        return 3
    elif 61 <= popularity <= 75:
        return 2
    elif 50 <= popularity <= 60:
        return 1
    return 0


def score_overall_popularity(popularity):
    """Score based on overall Search Popularity (1-100)."""
    if 86 <= popularity <= 100:
        return 5
    elif 71 <= popularity <= 85:
        return 4
    elif 61 <= popularity <= 70:
        return 3
    elif 50 <= popularity <= 60:
        return 2
    return 0


def process_from_database(country_filter="United States", report_id=None):
    """
    Read keywords from database and return scored keywords.

    Args:
        country_filter: Filter for this country
        report_id: Specific report ID to use, or None for most recent active report

    Returns:
        List of keyword dictionaries
    """
    # Get the report
    if report_id:
        report = execute_one(
            "SELECT * FROM apple_reports WHERE id = ?",
            (report_id,)
        )
        if not report:
            raise ValueError(f"Report ID {report_id} not found")
    else:
        # Get most recent active report for this country
        # Note: We filter by checking if any keywords exist for this country in the report
        report = execute_one(
            """SELECT DISTINCT ar.*
               FROM apple_reports ar
               JOIN apple_keywords ak ON ar.id = ak.report_id
               WHERE ar.is_active = 1 AND ak.country = ?
               ORDER BY ar.generated_at DESC
               LIMIT 1""",
            (country_filter,)
        )

        if not report:
            raise ValueError(
                f"No active reports found for country: {country_filter}. "
                f"Import a report with: python3 commands/import_report.py <excel_file>"
            )

    report_id = report['id']
    data_month = report['data_month']
    user_locale = report['user_locale']

    print(f"Reading from database...", file=sys.stderr)
    print(f"  Report ID: {report_id} ({report['report_id']})", file=sys.stderr)
    print(f"  Data Month: {data_month}", file=sys.stderr)
    print(f"  Locale: {user_locale}", file=sys.stderr)
    print(f"  Country filter: {country_filter}", file=sys.stderr)

    # Query keywords for this report and country
    rows = execute_query(
        """SELECT month.value as month, country, genre, search_term,
                  rank_in_genre, popularity_genre, popularity_overall,
                  score_rank, score_genre, score_overall, total_score
           FROM apple_keywords ak
           JOIN (SELECT ? as value) month
           WHERE ak.report_id = ? AND ak.country = ?
           ORDER BY ak.total_score DESC""",
        (data_month, report_id, country_filter)
    )

    if not rows:
        print(f"Warning: No keywords found for {country_filter}", file=sys.stderr)
        return []

    keywords = [dict(row) for row in rows]

    print(f"Retrieved {len(keywords)} keywords", file=sys.stderr)

    return keywords


def process_excel(excel_path, country_filter="United States"):
    """
    Process the Excel file and return scored keywords.

    Expected columns:
    - Month
    - Country or Region
    - Genre
    - Search Term
    - Rank in Genre
    - Search Popularity in Genre (1-100)
    - Search Popularity (1-100)
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active

    # Get header row (row 7 in this specific format)
    # Skip first 6 rows and get row 7
    headers = None
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 7:
            headers = list(row) if row else None
            # Validate we found the right row
            if headers and headers[0] == "Month":
                break
            headers = None

    if not headers:
        raise ValueError(f"Could not find valid header row at row 7. Check file format.")

    # Find column indices
    try:
        month_idx = headers.index("Month")
        country_idx = headers.index("Country or Region")
        genre_idx = headers.index("Genre")
        term_idx = headers.index("Search Term")
        rank_idx = headers.index("Rank in Genre")
        pop_genre_idx = headers.index("Search Popularity in Genre (1-100)")
        pop_overall_idx = headers.index("Search Popularity (1-100)")
    except ValueError as e:
        print(f"Error: Could not find required column. Headers: {headers}", file=sys.stderr)
        raise

    keywords = []
    processed_count = 0
    skipped_count = 0

    # Process rows (start from row 8, after metadata and header) using iter_rows for efficiency
    for row in sheet.iter_rows(min_row=8, values_only=True):
        processed_count += 1

        # Progress indicator every 10k rows
        if processed_count % 10000 == 0:
            print(f"Processed {processed_count} rows...", file=sys.stderr)

        if not row[term_idx]:  # Skip empty search terms
            skipped_count += 1
            continue

        # Filter by country if specified
        country = row[country_idx]
        if country_filter and country != country_filter:
            skipped_count += 1
            continue

        try:
            rank = int(row[rank_idx]) if row[rank_idx] else 999
            pop_genre = int(row[pop_genre_idx]) if row[pop_genre_idx] else 0
            pop_overall = int(row[pop_overall_idx]) if row[pop_overall_idx] else 0
        except (ValueError, TypeError):
            skipped_count += 1
            continue

        # Calculate scores
        rank_score = score_rank_in_genre(rank)
        genre_score = score_popularity_in_genre(pop_genre)
        overall_score = score_overall_popularity(pop_overall)
        total_score = rank_score + genre_score + overall_score

        keywords.append({
            "month": str(row[month_idx]) if row[month_idx] else "",
            "country": country,
            "genre": str(row[genre_idx]) if row[genre_idx] else "",
            "search_term": str(row[term_idx]),
            "rank_in_genre": rank,
            "popularity_genre": pop_genre,
            "popularity_overall": pop_overall,
            "score_rank": rank_score,
            "score_genre": genre_score,
            "score_overall": overall_score,
            "total_score": total_score
        })

    print(f"Processed {processed_count} total rows, skipped {skipped_count} rows", file=sys.stderr)

    wb.close()

    # Sort by total score (descending)
    keywords.sort(key=lambda x: x["total_score"], reverse=True)

    return keywords


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Process Apple Search Ads keyword rankings",
        epilog="By default, reads from database. Use --from-excel to read from Excel file."
    )
    parser.add_argument(
        "--country",
        default="United States",
        help="Country to filter for (default: United States)"
    )
    parser.add_argument(
        "--report-id",
        type=int,
        help="Specific report ID to use (default: most recent active report)"
    )
    parser.add_argument(
        "--from-excel",
        metavar="EXCEL_FILE",
        help="Read from Excel file instead of database (legacy mode)"
    )

    args = parser.parse_args()

    try:
        if args.from_excel:
            # Legacy mode: read from Excel
            if openpyxl is None:
                print("Error: openpyxl not installed. Install with: pip3 install openpyxl", file=sys.stderr)
                sys.exit(1)

            excel_path = Path(args.from_excel)
            if not excel_path.exists():
                print(f"Error: File not found: {excel_path}", file=sys.stderr)
                sys.exit(1)

            print(f"Processing {excel_path}...", file=sys.stderr)
            print(f"Filtering for country: {args.country}", file=sys.stderr)

            keywords = process_excel(excel_path, args.country)
        else:
            # Default mode: read from database
            keywords = process_from_database(
                country_filter=args.country,
                report_id=args.report_id
            )

        # Output JSON to stdout
        output = {
            "country": args.country,
            "total_keywords": len(keywords),
            "keywords": keywords
        }

        print(json.dumps(output, indent=2))

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
